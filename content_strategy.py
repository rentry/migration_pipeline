"""
content_strategy.py

Reads a Content Strategy worksheet (.xlsx) and produces a structured list of
pages with a resolved copy status (existing / new / needs-review).

Design constraints this module exists to satisfy:
- Column ORDER varies sheet-to-sheet and project-to-project. Columns are
  resolved by HEADER TEXT, never by position.
- Missing expected headers must fail with a clear, specific error naming
  the sheet and the missing header -- never a silent misread or a bare
  IndexError/KeyError.
- "No existing copy" is not a single case. A page can be:
    * a normal single-URL page (existing copy to migrate)
    * explicitly new, in several inconsistent phrasings
      ("New page", "New section", "New landing page", or a full sentence
      containing "new page")
    * a cross-link / duplicate reference to another IA page, not a real
      standalone page at all
    * multiple comma-separated source URLs (content needs synthesis)
    * simply blank with no explanation -- genuinely ambiguous, must be
      flagged for human review, never guessed at
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Configuration: expected headers
# ---------------------------------------------------------------------------

# Headers we require to do our job. If any of these is missing from a sheet,
# we raise a clear error rather than guessing at a column position.
REQUIRED_HEADERS = [
    "Page ID",
    "Page name",
    "Current source URL",
]

# Headers we use if present, but don't require.
OPTIONAL_HEADERS = [
    "Descriptive new URL",
    "Department tag",
    "Page content notes (optional)",
]

HEADER_ROW = 5      # row 5 holds column headers in this workbook's layout
DATA_START_ROW = 6  # page rows begin at row 6

# Phrases that mean "this page is intentionally new, no existing copy".
# Matched case-insensitively as a substring, since real-world entries range
# from a bare label to a full sentence.
NEW_CONTENT_PHRASES = [
    "new page",
    "new section",
    "new landing page",
]

# Pattern for cross-link / duplicate references disguised as source URLs,
# e.g. "Admitted Student Next Steps — Link to 2.2" or "(Crosslink to summer
# programs?)". These are navigation aliases, not real distinct pages.
CROSSLINK_PATTERN = re.compile(r"\b(link to|crosslink to|cross link to)\b", re.IGNORECASE)


class ContentStrategyError(ValueError):
    """Raised when the workbook doesn't match the expected structure."""


@dataclass
class Page:
    sheet: str
    page_id: str
    page_name: str
    source_url_raw: Optional[str]
    descriptive_new_url: Optional[str] = None
    department_tag: Optional[str] = None
    notes: Optional[str] = None

    # Populated by classify()
    status: str = ""            # 'existing' | 'new' | 'crosslink' | 'needs_review'
    source_urls: list[str] = field(default_factory=list)  # parsed, may be >1
    review_reason: Optional[str] = None

    def classify(self) -> None:
        """Determine copy status from source_url_raw. Mutates self."""
        raw = self.source_url_raw

        if raw is None or str(raw).strip() == "":
            # A blank source URL sometimes still carries a crosslink signal
            # in the page NAME itself (e.g. "Admitted Student Next Steps —
            # Link to 2.2"), rather than in the source-URL field. Check
            # there too before falling back to a generic "blank" flag.
            if CROSSLINK_PATTERN.search(self.page_name):
                self.status = "crosslink"
                self.review_reason = (
                    f"Page name suggests this is a navigation cross-link, "
                    f"not a standalone page: {self.page_name!r}"
                )
                return
            if self.department_tag:
                # Has a department tag but no direct URL -- program content
                # likely lives on a department page. Still worth a human
                # glance rather than a silent assumption.
                self.status = "needs_review"
                self.review_reason = (
                    "Blank source URL, but has a department tag "
                    f"({self.department_tag}); confirm whether content "
                    "lives on the department page or is genuinely missing."
                )
            else:
                self.status = "needs_review"
                self.review_reason = "Blank source URL with no explanation."
            return

        text = str(raw).strip()

        if CROSSLINK_PATTERN.search(text):
            self.status = "crosslink"
            self.review_reason = f"Looks like a navigation cross-link, not a standalone page: {text!r}"
            return

        if any(phrase in text.lower() for phrase in NEW_CONTENT_PHRASES):
            self.status = "new"
            return

        # Split on commas for the multi-URL case, keep only things that look
        # like URLs. If nothing parses as a URL, treat as needs_review rather
        # than silently accepting prose as a "source".
        candidates = [c.strip() for c in text.split(",")]
        urls = [c for c in candidates if c.lower().startswith("http")]

        if not urls:
            self.status = "needs_review"
            self.review_reason = f"Source URL field is neither a URL, a recognized 'new' phrase, nor a cross-link: {text!r}"
            return

        self.source_urls = urls
        self.status = "existing"
        if len(urls) > 1:
            self.review_reason = (
                f"{len(urls)} source URLs listed for one page; "
                "content will need to be synthesized from multiple sources."
            )


def _resolve_headers(sheet_name: str, header_row_values: list) -> dict[str, int]:
    """Map header text -> column index (0-based) for one sheet.

    Raises ContentStrategyError naming the sheet and the missing header(s)
    if any REQUIRED_HEADERS are absent.
    """
    header_map: dict[str, int] = {}
    for idx, val in enumerate(header_row_values):
        if val is None:
            continue
        header_map[str(val).strip()] = idx

    missing = [h for h in REQUIRED_HEADERS if h not in header_map]
    if missing:
        raise ContentStrategyError(
            f"Sheet '{sheet_name}' is missing required header(s): {missing}. "
            f"Found headers: {[v for v in header_row_values if v is not None]}. "
            f"Expected row {HEADER_ROW} to contain these column titles "
            "(order doesn't matter, but the text must match exactly)."
        )
    return header_map


def read_content_strategy(path: str | Path) -> list[Page]:
    """Read every page row from every sheet in the workbook.

    Returns a flat list of Page objects, each already classified.
    Raises ContentStrategyError if any sheet is missing required headers.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Content strategy workbook not found: {path}")

    wb = load_workbook(path, data_only=True)
    pages: list[Page] = []
    errors: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        header_row_values = [
            c.value for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, max_col=40))
        ]

        try:
            headers = _resolve_headers(sheet_name, header_row_values)
        except ContentStrategyError as e:
            # Collect errors across all sheets before failing, so one run
            # tells you about every problem sheet, not just the first.
            errors.append(str(e))
            continue

        idx_pid = headers["Page ID"]
        idx_name = headers["Page name"]
        idx_src = headers["Current source URL"]
        idx_new_url = headers.get("Descriptive new URL")
        idx_dept = headers.get("Department tag")
        idx_notes = headers.get("Page content notes (optional)")

        for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, max_col=40):
            page_id = row[idx_pid].value
            if page_id is None:
                continue  # blank spacer row

            page = Page(
                sheet=sheet_name,
                page_id=str(page_id).strip(),
                page_name=(row[idx_name].value or "").strip() if row[idx_name].value else "",
                source_url_raw=row[idx_src].value,
                descriptive_new_url=(row[idx_new_url].value if idx_new_url is not None else None),
                department_tag=(row[idx_dept].value if idx_dept is not None else None),
                notes=(row[idx_notes].value if idx_notes is not None else None),
            )
            page.classify()
            pages.append(page)

    if errors:
        raise ContentStrategyError(
            f"{len(errors)} sheet(s) failed header validation:\n\n" + "\n\n".join(errors)
        )

    return pages


def summarize(pages: list[Page]) -> dict:
    """Quick counts by status, for a sanity-check report before extraction runs."""
    from collections import Counter
    counts = Counter(p.status for p in pages)
    return {
        "total_pages": len(pages),
        "existing": counts.get("existing", 0),
        "new": counts.get("new", 0),
        "crosslink": counts.get("crosslink", 0),
        "needs_review": counts.get("needs_review", 0),
        "multi_source_pages": [p.page_id for p in pages if len(p.source_urls) > 1],
    }


def needs_review_report(pages: list[Page]) -> list[dict]:
    """Flat list suitable for writing straight to a 'needs review' sheet/CSV."""
    return [
        {
            "sheet": p.sheet,
            "page_id": p.page_id,
            "page_name": p.page_name,
            "status": p.status,
            "source_url_raw": p.source_url_raw,
            "reason": p.review_reason,
        }
        for p in pages
        if p.status in ("needs_review", "crosslink") or p.review_reason
    ]


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) != 2:
        print("Usage: python content_strategy.py <path-to-content-strategy.xlsx>")
        sys.exit(1)

    pages = read_content_strategy(sys.argv[1])
    print(json.dumps(summarize(pages), indent=2))
    print("\n--- Needs review / flagged ---")
    for row in needs_review_report(pages):
        print(f"[{row['sheet']}] {row['page_id']} — {row['page_name']!r}: {row['reason']}")
